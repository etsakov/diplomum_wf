import requests
import json
from datetime import datetime
from openpyxl import Workbook
from optparse import OptionParser

def connect_to_stream():

    domainDict = { 'live' : 'stream-fxtrade.oanda.com',
               'demo' : 'stream-fxpractice.oanda.com' }

    # Replace the following variables with your personal values 
    environment = "demo" # Replace this 'live' if you wish to connect to the live environment 
    domain = domainDict[environment] 
    access_token = 'd612e5b6e0f902b605207883dded35bc-8928c64998a13c444885331fd77cf935'
    account_id = '101-004-6259640-001'
    instruments = 'EUR_USD' 

    try:
        s = requests.Session()
        url = 'https://' + domain + '/v3/accounts/101-004-6259640-001/pricing/stream'
        headers = {'Authorization' : 'Bearer ' + access_token,
                   # 'X-Accept-Datetime-Format' : 'unix'
                  }
        params = {'instruments' : instruments, 'accountId' : account_id}
        req = requests.Request('GET', url, headers = headers, params = params)
        pre = req.prepare()
        resp = s.send(pre, stream = True, verify = True)
        return resp
    except Exception as e:
        s.close()
        print("Caught exception when connecting to stream\n" + str(e)) 

# def write_price(cur_states)...

def demo(displayHeartbeat):
    response = connect_to_stream()
    if response.status_code != 200:
        print(response.text)
        return
    states_collection = list()
    item = 1

    # current_date = datetime.now().strftime('%d-%m-%Y')
    # workbook = Workbook()
    # dest_filename = str(current_date) + '.xlsx'
    # worksheet1 = workbook.active
    # worksheet1.title = 'Data log'
    # worksheet1.cell(row = 1, column = 1).value = 'date_time'
    # worksheet1.cell(row = 1, column = 2).value = 'status'
    # worksheet1.cell(row = 1, column = 3).value = 'ask_closeout'
    # worksheet1.cell(row = 1, column = 4).value = 'bid_closeout'
    # worksheet1.cell(row = 1, column = 5).value = 'ask'
    # worksheet1.cell(row = 1, column = 6).value = 'bid'
    # worksheet1.cell(row = 1, column = 7).value = 'spread'

    for line in response.iter_lines(1):
        if line:
            try:
                line = line.decode('utf-8')
                msg = json.loads(line)
            except Exception as e:
                print("Caught exception when converting message into json\n" + str(e))
                return

            if "instrument" in msg or "tick" in msg or displayHeartbeat:
                
                print("item =", item)
                print(msg)

                cur_state = dict()
                
                if msg['type'] == 'PRICE':
                    cur_state['date_time'] = datetime.strftime(datetime.strptime(msg['time'].split('.')[0], '%Y-%m-%dT%H:%M:%S'), '%d.%m.%Y %H:%M:%S')
                    cur_state['status'] = msg['status']
                    cur_state['instruments'] = msg['instrument']
                    cur_state['ask_closeout'] = msg['closeoutAsk']
                    cur_state['bid_closeout'] = msg['closeoutBid']
                    cur_state['ask'] = msg['asks'][0]['price']
                    cur_state['bid'] = msg['bids'][0]['price']
                    cur_state['spread'] = ('%.5f' % (float(msg['asks'][0]['price']) - float(msg['bids'][0]['price'])))

                else:
                    replacement = len(states_collection) - 1
                    cur_state['date_time'] = datetime.strftime(datetime.strptime(msg['time'].split('.')[0], '%Y-%m-%dT%H:%M:%S'), '%d.%m.%Y %H:%M:%S')
                    cur_state['status'] = 'HEARTBEAT'
                    cur_state['ask_closeout'] = states_collection[replacement]['ask_closeout']
                    cur_state['bid_closeout'] = states_collection[replacement]['bid_closeout']
                    cur_state['ask'] = states_collection[replacement]['ask']
                    cur_state['bid'] = states_collection[replacement]['bid']
                    cur_state['spread'] = states_collection[replacement]['spread']
                # write_price(cur_states)
                # Here you can insert any handler
                print(cur_state)
                states_collection.append(cur_state)

                if len(states_collection) < 15:
                    pass
                else:
                    print('TO DELETE:', states_collection[-15]['date_time'])
                    del(states_collection[-15])
                print('len(states_collection)', len(states_collection))

                # save_state = item % 12
                # ocup_lines = 1
                # if save_state == 0:
                #         row = ocup_lines + 1
                #         print('row', type(row), ' ', row)
                #         iter_row = len(states_collection) - 12
                #         print('iter_row', type(iter_row), ' ', iter_row)
                #         for data_row in states_collection:
                #             worksheet1.cell(row = row, column = 1).value = states_collection[iter_row]['date_time']
                #             worksheet1.cell(row = row, column = 2).value = states_collection[iter_row]['status']
                #             worksheet1.cell(row = row, column = 3).value = states_collection[iter_row]['ask_closeout']
                #             worksheet1.cell(row = row, column = 4).value = states_collection[iter_row]['bid_closeout']
                #             worksheet1.cell(row = row, column = 5).value = states_collection[iter_row]['ask']
                #             worksheet1.cell(row = row, column = 6).value = states_collection[iter_row]['bid']
                #             worksheet1.cell(row = row, column = 7).value = states_collection[iter_row]['spread']
                #             iter_row += 1
                #             row += 1

                #         workbook.save(filename=dest_filename)
                #         print('SAVED')
                #         ocup_lines += 12
                # else:
                #         pass

                item += 1


# TO DO - Learn how to add the new data but not to re-write file

# NB: since "save_state = row % 12" the file to be saved once a minute


# OptionParser parses the command line

def main():
    usage = "usage: %prog [options]"
    parser = OptionParser(usage)
    parser.add_option("-b", "--displayHeartBeat", dest = "verbose", action = "store_true", 
                        help = "Display HeartBeat in streaming data")
    displayHeartbeat = False

    (options, args) = parser.parse_args()
    if len(args) > 1:
        parser.error("incorrect number of arguments")
    if options.verbose:
        displayHeartbeat = True
    demo(displayHeartbeat)


if __name__ == "__main__":
    main()